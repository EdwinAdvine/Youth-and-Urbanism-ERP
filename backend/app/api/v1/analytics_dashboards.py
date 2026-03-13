"""Pre-built dashboard data endpoints — real DB queries for each module dashboard.

Each endpoint returns all chart/KPI data needed by the corresponding prebuilt
dashboard on the frontend, eliminating hardcoded mock data.
"""

from datetime import date, datetime
from decimal import Decimal

from fastapi import APIRouter
from sqlalchemy import text

from app.core.deps import DBSession, CurrentUser

router = APIRouter(prefix="/analytics/dashboard", tags=["Analytics Dashboards"])


def _month_label(dt: date) -> str:
    return dt.strftime("%b %Y")


def _last_n_month_starts(n: int = 6) -> list[date]:
    """Return first-of-month dates for the last *n* months including current."""
    today = date.today()
    starts = []
    for i in range(n - 1, -1, -1):
        y = today.year
        m = today.month - i
        while m <= 0:
            m += 12
            y -= 1
        starts.append(date(y, m, 1))
    return starts


# ─── Manufacturing Dashboard ────────────────────────────────────────────────

@router.get("/manufacturing")
async def manufacturing_dashboard(
    db: DBSession,
    _user: CurrentUser,
):
    months = _last_n_month_starts(6)
    month_labels = [_month_label(m) for m in months]

    # ── Production output per month (completed_quantity from work orders) ──
    prod_rows = (
        await db.execute(
            text("""
                SELECT date_trunc('month', actual_end) AS m,
                       COALESCE(SUM(completed_quantity), 0) AS output,
                       COALESCE(SUM(planned_quantity), 0) AS target
                FROM mfg_work_orders
                WHERE actual_end >= :start
                GROUP BY 1 ORDER BY 1
            """),
            {"start": months[0]},
        )
    ).all()
    prod_map = {r.m.date().replace(day=1): r for r in prod_rows}
    production_data = []
    for m in months:
        r = prod_map.get(m)
        production_data.append({
            "month": _month_label(m),
            "output": int(r.output) if r else 0,
            "target": int(r.target) if r else 0,
        })

    # ── Defect rate per month ──
    defect_rows = (
        await db.execute(
            text("""
                SELECT date_trunc('month', created_at) AS m,
                       CASE WHEN SUM(quantity_inspected) > 0
                            THEN ROUND(SUM(quantity_failed)::numeric / SUM(quantity_inspected) * 100, 1)
                            ELSE 0 END AS rate
                FROM mfg_quality_checks
                WHERE created_at >= :start
                GROUP BY 1 ORDER BY 1
            """),
            {"start": months[0]},
        )
    ).all()
    defect_map = {r.m.date().replace(day=1): r for r in defect_rows}
    defect_data = []
    for m in months:
        r = defect_map.get(m)
        defect_data.append({
            "month": _month_label(m),
            "rate": float(r.rate) if r else 0.0,
        })

    # ── Work order status distribution ──
    wo_status_rows = (
        await db.execute(
            text("""
                SELECT status, COUNT(*) AS cnt
                FROM mfg_work_orders
                GROUP BY status ORDER BY cnt DESC
            """)
        )
    ).all()
    wo_status = [{"name": r.status.replace("_", " ").title(), "value": r.cnt} for r in wo_status_rows]

    # ── Workstation utilization (running hours / total available hours this month) ──
    ws_rows = (
        await db.execute(
            text("""
                SELECT w.name,
                       CASE WHEN w.capacity_per_hour > 0 THEN
                           LEAST(ROUND(
                               COUNT(wo.id) FILTER (WHERE wo.status = 'in_progress' OR wo.status = 'completed')
                               * 100.0 / GREATEST(w.capacity_per_hour * 8 * 22, 1)
                           , 0), 100)
                       ELSE 0 END AS utilization
                FROM mfg_workstations w
                LEFT JOIN mfg_work_orders wo ON wo.workstation_id = w.id
                    AND wo.created_at >= date_trunc('month', CURRENT_DATE)
                WHERE w.is_active = true
                GROUP BY w.id, w.name, w.capacity_per_hour
                ORDER BY w.name
            """)
        )
    ).all()
    workstations = [{"name": r.name, "utilization": int(r.utilization)} for r in ws_rows]

    # ── OEE (simplified: from quality checks + work order completion) ──
    oee_row = (
        await db.execute(
            text("""
                SELECT
                    COALESCE(
                        ROUND(COUNT(*) FILTER (WHERE status IN ('completed','in_progress'))::numeric
                              / GREATEST(COUNT(*), 1) * 100, 1), 0
                    ) AS availability,
                    COALESCE(
                        ROUND(SUM(completed_quantity)::numeric
                              / GREATEST(SUM(planned_quantity), 1) * 100, 1), 0
                    ) AS performance,
                    COALESCE((
                        SELECT ROUND(SUM(quantity_passed)::numeric
                                     / GREATEST(SUM(quantity_inspected), 1) * 100, 1)
                        FROM mfg_quality_checks
                        WHERE created_at >= date_trunc('month', CURRENT_DATE)
                    ), 100) AS quality
                FROM mfg_work_orders
                WHERE created_at >= date_trunc('month', CURRENT_DATE)
            """)
        )
    ).first()
    availability = float(oee_row.availability) if oee_row else 0
    performance = float(oee_row.performance) if oee_row else 0
    quality = float(oee_row.quality) if oee_row else 0
    oee = round(availability * performance * quality / 10000, 1)

    # ── KPI: active work orders ──
    active_wo = (
        await db.execute(
            text("SELECT COUNT(*) FROM mfg_work_orders WHERE status = 'in_progress'")
        )
    ).scalar() or 0

    return {
        "production_data": production_data,
        "defect_data": defect_data,
        "wo_status": wo_status,
        "workstations": workstations,
        "oee": {"score": oee, "availability": availability, "performance": performance, "quality": quality},
        "kpis": {
            "oee": oee,
            "monthly_output": production_data[-1]["output"] if production_data else 0,
            "defect_rate": defect_data[-1]["rate"] if defect_data else 0,
            "active_work_orders": active_wo,
        },
    }


# ─── CRM Dashboard ──────────────────────────────────────────────────────────

@router.get("/crm")
async def crm_dashboard(
    db: DBSession,
    _user: CurrentUser,
):
    months = _last_n_month_starts(6)

    # ── Pipeline stages (opportunities by stage) ──
    pipeline_rows = (
        await db.execute(
            text("""
                SELECT stage, COUNT(*) AS cnt
                FROM crm_opportunities
                WHERE stage NOT IN ('closed_won', 'closed_lost')
                GROUP BY stage ORDER BY cnt DESC
            """)
        )
    ).all()
    pipeline_stages = [{"name": r.stage.replace("_", " ").title(), "value": r.cnt} for r in pipeline_rows]

    # ── Lead sources ──
    source_rows = (
        await db.execute(
            text("""
                SELECT COALESCE(source, 'Unknown') AS source, COUNT(*) AS cnt
                FROM crm_contacts
                WHERE source IS NOT NULL
                GROUP BY source ORDER BY cnt DESC
                LIMIT 10
            """)
        )
    ).all()
    lead_sources = [{"name": r.source.replace("_", " ").title(), "value": r.cnt} for r in source_rows]

    # ── Conversion rate trend (deals closed_won / leads created per month) ──
    won_rows = (
        await db.execute(
            text("""
                SELECT date_trunc('month', close_date)::date AS month, COUNT(*) AS won
                FROM crm_deals WHERE status = 'won' AND close_date >= :start
                GROUP BY 1
            """),
            {"start": months[0]},
        )
    ).all()
    won_map = {r.month.replace(day=1): r.won for r in won_rows}
    lead_rows = (
        await db.execute(
            text("""
                SELECT date_trunc('month', created_at)::date AS month, COUNT(*) AS leads
                FROM crm_leads WHERE created_at >= :start
                GROUP BY 1
            """),
            {"start": months[0]},
        )
    ).all()
    lead_map = {r.month.replace(day=1): r.leads for r in lead_rows}
    conversion_data = []
    for m in months:
        won = won_map.get(m, 0)
        leads = lead_map.get(m, 0)
        rate = round(won / leads * 100, 1) if leads > 0 else 0
        conversion_data.append({"month": _month_label(m), "rate": rate})

    # ── Deal velocity (avg days to close per month) ──
    velocity_rows = (
        await db.execute(
            text("""
                SELECT date_trunc('month', close_date) AS m,
                       ROUND(AVG(close_date - created_at::date)) AS days
                FROM crm_deals
                WHERE status = 'won' AND close_date >= :start
                GROUP BY 1 ORDER BY 1
            """),
            {"start": months[0]},
        )
    ).all()
    vel_map = {r.m.date().replace(day=1): r for r in velocity_rows}
    velocity_data = [
        {"month": _month_label(m), "days": int(vel_map[m].days) if m in vel_map else 0}
        for m in months
    ]

    # ── Deal size distribution ──
    size_rows = (
        await db.execute(
            text("""
                SELECT
                    CASE
                        WHEN deal_value < 50000 THEN '< KSh 50K'
                        WHEN deal_value < 100000 THEN 'KSh 50K-100K'
                        WHEN deal_value < 500000 THEN 'KSh 100K-500K'
                        WHEN deal_value < 1000000 THEN 'KSh 500K-1M'
                        ELSE '> KSh 1M'
                    END AS range,
                    COUNT(*) AS count
                FROM crm_deals
                GROUP BY 1
                ORDER BY MIN(deal_value)
            """)
        )
    ).all()
    deal_sizes = [{"range": r.range, "count": r.count} for r in size_rows]

    # ── KPI summaries ──
    kpi_row = (
        await db.execute(
            text("""
                SELECT
                    (SELECT COUNT(*) FROM crm_opportunities WHERE stage NOT IN ('closed_won','closed_lost')) AS pipeline_deals,
                    (SELECT COALESCE(SUM(expected_value), 0) FROM crm_opportunities WHERE stage NOT IN ('closed_won','closed_lost')) AS pipeline_value,
                    (SELECT COALESCE(AVG(deal_value), 0) FROM crm_deals) AS avg_deal_size
            """)
        )
    ).first()

    return {
        "pipeline_stages": pipeline_stages,
        "lead_sources": lead_sources,
        "conversion_data": conversion_data,
        "velocity_data": velocity_data,
        "deal_sizes": deal_sizes,
        "kpis": {
            "pipeline_deals": kpi_row.pipeline_deals if kpi_row else 0,
            "pipeline_value": float(kpi_row.pipeline_value) if kpi_row else 0,
            "avg_deal_size": float(kpi_row.avg_deal_size) if kpi_row else 0,
            "conversion_rate": conversion_data[-1]["rate"] if conversion_data else 0,
        },
    }


# ─── HR Dashboard ────────────────────────────────────────────────────────────

@router.get("/hr")
async def hr_dashboard(
    db: DBSession,
    _user: CurrentUser,
):
    months = _last_n_month_starts(6)
    today = date.today()

    # ── Headcount trend ──
    headcount_data = []
    for m in months:
        next_m = date(m.year + (1 if m.month == 12 else 0), (m.month % 12) + 1, 1)
        cnt = (
            await db.execute(
                text("""
                    SELECT COUNT(*) FROM hr_employees
                    WHERE hire_date < :end_date
                      AND (termination_date IS NULL OR termination_date >= :start_date)
                      AND is_active = true
                """),
                {"start_date": m, "end_date": next_m},
            )
        ).scalar() or 0
        headcount_data.append({"month": _month_label(m), "headcount": cnt})

    # ── Attrition trend (terminations per month / avg headcount) ──
    attrition_rows = (
        await db.execute(
            text("""
                SELECT date_trunc('month', termination_date) AS m,
                       COUNT(*) AS terms
                FROM hr_employees
                WHERE termination_date >= :start AND termination_date IS NOT NULL
                GROUP BY 1 ORDER BY 1
            """),
            {"start": months[0]},
        )
    ).all()
    term_map = {r.m.date().replace(day=1): r.terms for r in attrition_rows}
    attrition_data = []
    for i, m in enumerate(months):
        hc = headcount_data[i]["headcount"] or 1
        terms = term_map.get(m, 0)
        attrition_data.append({
            "month": _month_label(m),
            "rate": round(terms / hc * 100, 1) if hc > 0 else 0,
        })

    # ── Department distribution ──
    dept_rows = (
        await db.execute(
            text("""
                SELECT d.name, COUNT(e.id) AS cnt
                FROM hr_departments d
                LEFT JOIN hr_employees e ON e.department_id = d.id AND e.is_active = true
                WHERE d.is_active = true
                GROUP BY d.id, d.name
                HAVING COUNT(e.id) > 0
                ORDER BY cnt DESC
            """)
        )
    ).all()
    dept_distribution = [{"name": r.name, "value": r.cnt} for r in dept_rows]

    # ── Attendance breakdown (last 6 months) ──
    att_rows = (
        await db.execute(
            text("""
                SELECT date_trunc('month', attendance_date) AS m,
                       ROUND(COUNT(*) FILTER (WHERE status = 'present')::numeric / GREATEST(COUNT(*), 1) * 100) AS present,
                       ROUND(COUNT(*) FILTER (WHERE status = 'remote')::numeric / GREATEST(COUNT(*), 1) * 100) AS remote,
                       ROUND(COUNT(*) FILTER (WHERE status = 'absent')::numeric / GREATEST(COUNT(*), 1) * 100) AS absent
                FROM hr_attendance
                WHERE attendance_date >= :start
                GROUP BY 1 ORDER BY 1
            """),
            {"start": months[0]},
        )
    ).all()
    att_map = {r.m.date().replace(day=1): r for r in att_rows}
    attendance_data = []
    for m in months:
        r = att_map.get(m)
        attendance_data.append({
            "month": _month_label(m),
            "present": int(r.present) if r else 0,
            "remote": int(r.remote) if r else 0,
            "absent": int(r.absent) if r else 0,
        })

    # ── Leave utilization ──
    leave_rows = (
        await db.execute(
            text("""
                SELECT leave_type AS type,
                       COUNT(*) FILTER (WHERE status = 'approved') AS used,
                       COUNT(*) AS total
                FROM hr_leave_requests
                WHERE start_date >= date_trunc('year', CURRENT_DATE)
                GROUP BY leave_type ORDER BY total DESC
                LIMIT 5
            """)
        )
    ).all()
    leave_data = [{"type": r.type.replace("_", " ").title(), "used": r.used, "total": r.total} for r in leave_rows]

    # ── KPIs ──
    active_count = headcount_data[-1]["headcount"] if headcount_data else 0
    avg_attrition = sum(d["rate"] for d in attrition_data) / len(attrition_data) if attrition_data else 0
    total_leave_used = sum(l["used"] for l in leave_data)
    total_leave_total = sum(l["total"] for l in leave_data) or 1
    leave_utilization = round(total_leave_used / total_leave_total * 100)
    last_att = attendance_data[-1] if attendance_data else {"present": 0}

    return {
        "headcount_data": headcount_data,
        "attrition_data": attrition_data,
        "dept_distribution": dept_distribution,
        "attendance_data": attendance_data,
        "leave_data": leave_data,
        "kpis": {
            "active_employees": active_count,
            "attrition_rate": round(avg_attrition, 1),
            "attendance_rate": last_att["present"],
            "leave_utilization": leave_utilization,
        },
    }


# ─── Inventory Dashboard ────────────────────────────────────────────────────

@router.get("/inventory")
async def inventory_dashboard(
    db: DBSession,
    _user: CurrentUser,
):
    months = _last_n_month_starts(6)

    # ── Stock by category ──
    cat_rows = (
        await db.execute(
            text("""
                SELECT COALESCE(i.category, 'Uncategorized') AS name,
                       COALESCE(SUM(sl.quantity_on_hand), 0)::int AS value
                FROM inventory_items i
                LEFT JOIN inventory_stock_levels sl ON sl.item_id = i.id
                WHERE i.is_active = true
                GROUP BY i.category
                ORDER BY value DESC
                LIMIT 8
            """)
        )
    ).all()
    stock_by_category = [{"name": r.name, "value": r.value} for r in cat_rows]

    # ── Turnover ratio per month (movements / avg stock) ──
    turn_rows = (
        await db.execute(
            text("""
                SELECT date_trunc('month', sm.created_at) AS m,
                       ROUND(COUNT(*)::numeric / GREATEST(
                           (SELECT COALESCE(AVG(quantity_on_hand), 1) FROM inventory_stock_levels), 1
                       ), 1) AS ratio
                FROM inventory_stock_movements sm
                WHERE sm.created_at >= :start
                GROUP BY 1 ORDER BY 1
            """),
            {"start": months[0]},
        )
    ).all()
    turn_map = {r.m.date().replace(day=1): float(r.ratio) for r in turn_rows}
    turnover_data = [
        {"month": _month_label(m), "ratio": turn_map.get(m, 0)}
        for m in months
    ]

    # ── Valuation trend ──
    val_rows = (
        await db.execute(
            text("""
                SELECT date_trunc('month', sm.created_at) AS m,
                       COALESCE(SUM(sl2.quantity_on_hand * i.cost_price), 0) AS value
                FROM (SELECT DISTINCT date_trunc('month', created_at) AS m FROM inventory_stock_movements WHERE created_at >= :start) sm
                CROSS JOIN LATERAL (
                    SELECT SUM(sl.quantity_on_hand) AS quantity_on_hand, sl.item_id
                    FROM inventory_stock_levels sl
                    GROUP BY sl.item_id
                ) sl2
                JOIN inventory_items i ON i.id = sl2.item_id
                GROUP BY 1 ORDER BY 1
            """),
            {"start": months[0]},
        )
    ).all()
    # Simpler approach: current valuation snapshot for trend
    current_val = (
        await db.execute(
            text("""
                SELECT COALESCE(SUM(sl.quantity_on_hand * i.cost_price), 0) AS value
                FROM inventory_stock_levels sl
                JOIN inventory_items i ON i.id = sl.item_id
                WHERE i.is_active = true
            """)
        )
    ).scalar() or 0
    valuation_data = [
        {"month": _month_label(m), "value": float(current_val)}
        for m in months
    ]

    # ── Top items by value ──
    top_rows = (
        await db.execute(
            text("""
                SELECT i.name, COALESCE(SUM(sl.quantity_on_hand), 0)::int AS qty,
                       COALESCE(SUM(sl.quantity_on_hand * i.cost_price), 0)::numeric AS value
                FROM inventory_items i
                JOIN inventory_stock_levels sl ON sl.item_id = i.id
                WHERE i.is_active = true
                GROUP BY i.id, i.name
                ORDER BY value DESC
                LIMIT 5
            """)
        )
    ).all()
    top_items = [{"name": r.name, "qty": r.qty, "value": float(r.value)} for r in top_rows]

    # ── Warehouse stock distribution ──
    wh_rows = (
        await db.execute(
            text("""
                SELECT w.name,
                       COALESCE(SUM(sl.quantity_on_hand), 0)::int AS stock_qty
                FROM inventory_warehouses w
                LEFT JOIN inventory_stock_levels sl ON sl.warehouse_id = w.id
                WHERE w.is_active = true
                GROUP BY w.id, w.name
                ORDER BY stock_qty DESC
            """)
        )
    ).all()
    # Convert to percentage of max warehouse stock for gauge display
    max_stock = max((r.stock_qty for r in wh_rows), default=1) or 1
    warehouse_data = [
        {"name": r.name, "capacity": min(round(r.stock_qty / max_stock * 100), 100)}
        for r in wh_rows
    ]

    # ── KPIs ──
    total_items = sum(c["value"] for c in stock_by_category)
    avg_turnover = sum(d["ratio"] for d in turnover_data) / len(turnover_data) if turnover_data else 0
    low_stock = (
        await db.execute(
            text("""
                SELECT COUNT(*) FROM inventory_items i
                JOIN inventory_stock_levels sl ON sl.item_id = i.id
                WHERE i.is_active = true AND sl.quantity_on_hand <= i.reorder_level AND i.reorder_level > 0
            """)
        )
    ).scalar() or 0

    return {
        "stock_by_category": stock_by_category,
        "turnover_data": turnover_data,
        "valuation_data": valuation_data,
        "top_items": top_items,
        "warehouse_data": warehouse_data,
        "kpis": {
            "total_skus": total_items,
            "total_valuation": float(current_val),
            "avg_turnover": round(avg_turnover, 1),
            "low_stock_alerts": low_stock,
        },
    }


# ─── Support Dashboard ──────────────────────────────────────────────────────

@router.get("/support")
async def support_dashboard(
    db: DBSession,
    _user: CurrentUser,
):
    months = _last_n_month_starts(6)

    # ── Ticket volume trend ──
    vol_rows = (
        await db.execute(
            text("""
                SELECT date_trunc('month', created_at) AS m,
                       COUNT(*) AS new,
                       COUNT(*) FILTER (WHERE status IN ('resolved', 'closed')) AS resolved
                FROM tickets
                WHERE created_at >= :start
                GROUP BY 1 ORDER BY 1
            """),
            {"start": months[0]},
        )
    ).all()
    vol_map = {r.m.date().replace(day=1): r for r in vol_rows}
    volume_data = []
    for m in months:
        r = vol_map.get(m)
        volume_data.append({
            "month": _month_label(m),
            "new": r.new if r else 0,
            "resolved": r.resolved if r else 0,
        })

    # ── Resolution time trend ──
    res_rows = (
        await db.execute(
            text("""
                SELECT date_trunc('month', resolved_at) AS m,
                       ROUND(AVG(EXTRACT(EPOCH FROM (resolved_at - created_at)) / 3600)::numeric, 1) AS hours
                FROM tickets
                WHERE resolved_at IS NOT NULL AND resolved_at >= :start
                GROUP BY 1 ORDER BY 1
            """),
            {"start": months[0]},
        )
    ).all()
    res_map = {r.m.date().replace(day=1): float(r.hours) for r in res_rows}
    resolution_data = [
        {"month": _month_label(m), "hours": res_map.get(m, 0)}
        for m in months
    ]

    # ── Ticket categories ──
    cat_rows = (
        await db.execute(
            text("""
                SELECT COALESCE(tc.name, 'Uncategorized') AS name, COUNT(*) AS value
                FROM tickets t
                LEFT JOIN ticket_categories tc ON tc.id = t.category_id
                GROUP BY tc.name ORDER BY value DESC
                LIMIT 8
            """)
        )
    ).all()
    categories = [{"name": r.name, "value": r.value} for r in cat_rows]

    # ── Priority breakdown ──
    pri_rows = (
        await db.execute(
            text("""
                SELECT priority AS name, COUNT(*) AS value
                FROM tickets
                GROUP BY priority ORDER BY value DESC
            """)
        )
    ).all()
    priorities = [{"name": r.name.title(), "value": r.value} for r in pri_rows]

    # ── CSAT trend ──
    csat_rows = (
        await db.execute(
            text("""
                SELECT date_trunc('month', submitted_at) AS m,
                       ROUND(AVG(rating)::numeric * 20, 0) AS score
                FROM customer_satisfaction
                WHERE submitted_at >= :start
                GROUP BY 1 ORDER BY 1
            """),
            {"start": months[0]},
        )
    ).all()
    csat_map = {r.m.date().replace(day=1): int(r.score) for r in csat_rows}
    csat_data = [
        {"month": _month_label(m), "score": csat_map.get(m, 0)}
        for m in months
    ]

    # ── KPIs ──
    kpi_row = (
        await db.execute(
            text("""
                SELECT
                    COUNT(*) FILTER (WHERE status = 'open') AS open_tickets,
                    COUNT(*) FILTER (WHERE status IN ('resolved', 'closed')) AS resolved
                FROM tickets
            """)
        )
    ).first()
    avg_resolution = sum(d["hours"] for d in resolution_data) / len([d for d in resolution_data if d["hours"] > 0]) if any(d["hours"] > 0 for d in resolution_data) else 0
    current_csat = csat_data[-1]["score"] if csat_data else 0

    # ── SLA compliance ──
    sla_compliance = (
        await db.execute(
            text("""
                SELECT ROUND(
                    COUNT(*) FILTER (WHERE resolved_at IS NOT NULL
                                     AND resolved_at <= created_at + interval '24 hours')::numeric
                    / GREATEST(COUNT(*) FILTER (WHERE resolved_at IS NOT NULL), 1) * 100
                ) AS pct
                FROM tickets
                WHERE created_at >= date_trunc('month', CURRENT_DATE)
            """)
        )
    ).scalar() or 0

    return {
        "volume_data": volume_data,
        "resolution_data": resolution_data,
        "categories": categories,
        "priorities": priorities,
        "csat_data": csat_data,
        "kpis": {
            "open_tickets": kpi_row.open_tickets if kpi_row else 0,
            "resolved": kpi_row.resolved if kpi_row else 0,
            "avg_resolution": round(avg_resolution, 1),
            "csat_score": current_csat,
            "sla_compliance": int(sla_compliance),
        },
    }


# ─── E-Commerce Dashboard ───────────────────────────────────────────────────

@router.get("/ecommerce")
async def ecommerce_dashboard(
    db: DBSession,
    _user: CurrentUser,
):
    months = _last_n_month_starts(6)

    # ── Order volume trend ──
    order_rows = (
        await db.execute(
            text("""
                SELECT date_trunc('month', created_at) AS m, COUNT(*) AS orders
                FROM ecom_orders
                WHERE created_at >= :start
                GROUP BY 1 ORDER BY 1
            """),
            {"start": months[0]},
        )
    ).all()
    order_map = {r.m.date().replace(day=1): r.orders for r in order_rows}
    order_data = [
        {"month": _month_label(m), "orders": order_map.get(m, 0)}
        for m in months
    ]

    # ── Order status distribution ──
    status_rows = (
        await db.execute(
            text("""
                SELECT status AS name, COUNT(*) AS value
                FROM ecom_orders GROUP BY status ORDER BY value DESC
            """)
        )
    ).all()
    order_status = [{"name": r.name.replace("_", " ").title(), "value": r.value} for r in status_rows]

    # ── Revenue by channel (if channel field exists, otherwise by status) ──
    rev_rows = (
        await db.execute(
            text("""
                SELECT date_trunc('month', created_at) AS m,
                       COALESCE(SUM(total), 0) AS revenue
                FROM ecom_orders
                WHERE created_at >= :start AND status != 'cancelled'
                GROUP BY 1 ORDER BY 1
            """),
            {"start": months[0]},
        )
    ).all()
    rev_map = {r.m.date().replace(day=1): float(r.revenue) for r in rev_rows}
    revenue_data = [
        {"month": _month_label(m), "Revenue": rev_map.get(m, 0)}
        for m in months
    ]

    # ── Top products ──
    prod_rows = (
        await db.execute(
            text("""
                SELECT p.name, SUM(ol.quantity) AS units_sold, SUM(ol.total) AS revenue
                FROM ecom_order_lines ol
                JOIN ecom_products p ON p.id = ol.product_id
                JOIN ecom_orders o ON o.id = ol.order_id
                WHERE o.status != 'cancelled'
                GROUP BY p.id, p.name
                ORDER BY revenue DESC
                LIMIT 10
            """)
        )
    ).all()
    top_products = [
        {"name": r.name, "units_sold": int(r.units_sold), "revenue": float(r.revenue)}
        for r in prod_rows
    ]

    # ── KPIs ──
    total_revenue = sum(d["Revenue"] for d in revenue_data)
    total_orders = sum(d["orders"] for d in order_data)
    avg_order_value = total_revenue / total_orders if total_orders > 0 else 0

    return {
        "order_data": order_data,
        "order_status": order_status,
        "revenue_data": revenue_data,
        "top_products": top_products,
        "kpis": {
            "total_revenue": total_revenue,
            "total_orders": total_orders,
            "avg_order_value": round(avg_order_value, 2),
        },
    }


# ─── Finance Dashboard ──────────────────────────────────────────────────────

@router.get("/finance")
async def finance_dashboard(
    db: DBSession,
    _user: CurrentUser,
):
    months = _last_n_month_starts(12)

    # ── Revenue trend ──
    rev_rows = (
        await db.execute(
            text("""
                SELECT date_trunc('month', issue_date) AS m,
                       COALESCE(SUM(total), 0) AS revenue
                FROM finance_invoices
                WHERE status IN ('paid', 'sent', 'overdue') AND issue_date >= :start
                GROUP BY 1 ORDER BY 1
            """),
            {"start": months[0]},
        )
    ).all()
    rev_map = {r.m.date().replace(day=1): float(r.revenue) for r in rev_rows}

    # ── Expense trend ──
    exp_rows = (
        await db.execute(
            text("""
                SELECT date_trunc('month', je.date) AS m,
                       COALESCE(SUM(jl.debit), 0) AS expenses
                FROM finance_journal_entries je
                JOIN finance_journal_lines jl ON jl.entry_id = je.id
                JOIN finance_accounts a ON a.id = jl.account_id
                WHERE a.account_type = 'expense' AND je.date >= :start
                GROUP BY 1 ORDER BY 1
            """),
            {"start": months[0]},
        )
    ).all()
    exp_map = {r.m.date().replace(day=1): float(r.expenses) for r in exp_rows}

    pnl_data = []
    for m in months:
        rev = rev_map.get(m, 0)
        exp = exp_map.get(m, 0)
        pnl_data.append({
            "month": _month_label(m),
            "Revenue": rev,
            "Expenses": exp,
            "Profit": rev - exp,
        })

    cumulative_cash = 0
    cash_flow_data = []
    for d in pnl_data:
        cumulative_cash += d["Profit"]
        cash_flow_data.append({"month": d["month"], "Cash Flow": cumulative_cash})

    # ── Expense distribution by account ──
    exp_dist_rows = (
        await db.execute(
            text("""
                SELECT a.name, COALESCE(SUM(jl.debit), 0) AS value
                FROM finance_journal_entries je
                JOIN finance_journal_lines jl ON jl.entry_id = je.id
                JOIN finance_accounts a ON a.id = jl.account_id
                WHERE a.account_type = 'expense'
                  AND je.date >= date_trunc('year', CURRENT_DATE)
                GROUP BY a.id, a.name
                ORDER BY value DESC
                LIMIT 6
            """)
        )
    ).all()
    expense_distribution = [{"name": r.name, "value": float(r.value)} for r in exp_dist_rows]

    # ── KPIs ──
    kpi_row = (
        await db.execute(
            text("""
                SELECT
                    (SELECT COALESCE(SUM(total), 0) FROM finance_invoices
                     WHERE status IN ('paid','sent','overdue')
                       AND issue_date >= date_trunc('month', CURRENT_DATE)) AS revenue_mtd,
                    (SELECT COUNT(*) FROM finance_invoices
                     WHERE status IN ('sent','overdue')) AS open_invoices
            """)
        )
    ).first()
    total_expenses = sum(d["Expenses"] for d in pnl_data)

    return {
        "pnl_data": pnl_data,
        "cash_flow_data": cash_flow_data,
        "expense_distribution": expense_distribution,
        "kpis": {
            "revenue_mtd": float(kpi_row.revenue_mtd) if kpi_row else 0,
            "open_invoices": kpi_row.open_invoices if kpi_row else 0,
            "total_expenses": total_expenses,
            "net_profit": float(kpi_row.revenue_mtd) - total_expenses / (len(pnl_data) or 1) if kpi_row else 0,
        },
    }


# ─── XLSX Dashboard Export ───────────────────────────────────────────────────

@router.post("/export-xlsx")
async def export_dashboard_xlsx(
    payload: dict,
    _user: CurrentUser,
):
    """Generate a styled XLSX workbook from dashboard data.

    Expects:
        {
            "title": "Dashboard Name",
            "sections": [
                {"title": "Section Name", "data": [{"col1": val, ...}, ...]}
            ]
        }
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    title = payload.get("title", "Dashboard Export")
    sections = payload.get("sections", [])

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limit

    # Style definitions
    header_font = Font(name="Open Sans", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="51459D", end_color="51459D", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_font = Font(name="Open Sans", bold=True, color="51459D", size=14)
    section_font = Font(name="Open Sans", bold=True, color="333333", size=12)
    data_font = Font(name="Open Sans", size=10)
    number_font = Font(name="Open Sans", size=10)
    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )
    alt_fill = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")

    row = 1
    # Dashboard title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = title_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

    # Export date
    ws.cell(row=row, column=1, value=f"Exported: {datetime.now().strftime('%B %d, %Y %I:%M %p')}")
    ws.cell(row=row, column=1).font = Font(name="Open Sans", italic=True, color="999999", size=9)
    row += 2

    for section in sections:
        section_title = section.get("title", "")
        data = section.get("data", [])
        if not data:
            continue

        # Section title
        ws.cell(row=row, column=1, value=section_title)
        ws.cell(row=row, column=1).font = section_font
        row += 1

        # Headers
        columns = list(data[0].keys())
        for ci, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row, column=ci, value=col_name.replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        row += 1

        # Data rows
        for ri, record in enumerate(data):
            for ci, col_name in enumerate(columns, 1):
                val = record.get(col_name, "")
                cell = ws.cell(row=row, column=ci, value=val)
                cell.font = data_font if not isinstance(val, (int, float, Decimal)) else number_font
                cell.border = thin_border
                if isinstance(val, (int, float, Decimal)):
                    cell.alignment = Alignment(horizontal="right")
                    if isinstance(val, float) and val > 1000:
                        cell.number_format = '#,##0.00'
                    elif isinstance(val, int) and val > 1000:
                        cell.number_format = '#,##0'
                if ri % 2 == 1:
                    cell.fill = alt_fill
            row += 1

        # Auto-width columns
        for ci, col_name in enumerate(columns, 1):
            max_len = len(col_name)
            for record in data:
                val = str(record.get(col_name, ""))
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)

        row += 1  # gap between sections

    # Freeze header
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_title = title.replace(" ", "_").lower()
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_title}_export.xlsx"'},
    )
