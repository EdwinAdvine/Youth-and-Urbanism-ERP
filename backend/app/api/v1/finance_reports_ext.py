"""Finance API — Extended Reports (Cash Flow, Trial Balance, Aged Receivables/Payables, KPIs)."""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from typing import Any

from fastapi import APIRouter, Query
from sqlalchemy import and_, func, select

from app.core.deps import CurrentUser, DBSession
from app.models.finance import (
    Account,
    Expense,
    FixedAsset,
    Invoice,
    JournalEntry,
    JournalLine,
    Payment,
    VendorBill,
)

router = APIRouter()


# ── Cash Flow Statement ──────────────────────────────────────────────────────

@router.get("/reports/cash-flow", summary="Cash flow statement for a date range")
async def cash_flow_statement(
    current_user: CurrentUser,
    db: DBSession,
    start_date: date = Query(..., description="Start date"),
    end_date: date = Query(..., description="End date"),
) -> dict[str, Any]:
    # Operating activities: payments received minus expenses/vendor payments
    # Inflows: completed payments received (linked to sales invoices)
    inflows_result = await db.execute(
        select(
            func.coalesce(func.sum(Payment.amount), 0).label("total"),
        )
        .where(
            and_(
                Payment.payment_date >= start_date,
                Payment.payment_date <= end_date,
                Payment.status == "completed",
                Payment.invoice_id.isnot(None),
            )
        )
    )
    cash_inflows = Decimal(str(inflows_result.scalar() or 0))

    # Outflows: vendor bill payments
    vendor_payments_result = await db.execute(
        select(
            func.coalesce(func.sum(VendorBill.total), 0).label("total"),
        )
        .where(
            and_(
                VendorBill.status == "paid",
                VendorBill.updated_at >= datetime.combine(start_date, datetime.min.time()),
                VendorBill.updated_at <= datetime.combine(end_date, datetime.max.time()),
            )
        )
    )
    vendor_outflows = Decimal(str(vendor_payments_result.scalar() or 0))

    # Expense reimbursements
    expense_result = await db.execute(
        select(
            func.coalesce(func.sum(Expense.amount), 0).label("total"),
        )
        .where(
            and_(
                Expense.status == "reimbursed",
                Expense.expense_date >= start_date,
                Expense.expense_date <= end_date,
            )
        )
    )
    expense_outflows = Decimal(str(expense_result.scalar() or 0))

    operating_net = cash_inflows - vendor_outflows - expense_outflows

    # Investing activities: fixed asset purchases in the period
    asset_purchases_result = await db.execute(
        select(
            func.coalesce(func.sum(FixedAsset.purchase_cost), 0).label("total"),
        )
        .where(
            and_(
                FixedAsset.purchase_date >= start_date,
                FixedAsset.purchase_date <= end_date,
            )
        )
    )
    asset_purchases = Decimal(str(asset_purchases_result.scalar() or 0))
    investing_net = -asset_purchases

    # Financing activities: from journal entries tagged as equity/liability
    financing_result = await db.execute(
        select(
            func.coalesce(func.sum(JournalLine.credit - JournalLine.debit), 0).label("net"),
        )
        .join(JournalEntry, JournalEntry.id == JournalLine.journal_entry_id)
        .join(Account, Account.id == JournalLine.account_id)
        .where(
            and_(
                JournalEntry.status == "posted",
                JournalEntry.entry_date >= start_date,
                JournalEntry.entry_date <= end_date,
                Account.account_type.in_(["equity", "liability"]),
            )
        )
    )
    financing_net = Decimal(str(financing_result.scalar() or 0))

    net_change = operating_net + investing_net + financing_net

    return {
        "period": {"start_date": str(start_date), "end_date": str(end_date)},
        "operating_activities": {
            "cash_inflows": str(cash_inflows),
            "vendor_payments": str(vendor_outflows),
            "expense_reimbursements": str(expense_outflows),
            "net": str(operating_net),
        },
        "investing_activities": {
            "asset_purchases": str(asset_purchases),
            "net": str(investing_net),
        },
        "financing_activities": {
            "net": str(financing_net),
        },
        "net_change_in_cash": str(net_change),
    }


# ── Trial Balance (extended — includes all accounts with zero balances) ──────

@router.get("/reports/trial-balance-ext", summary="Extended trial balance with all accounts")
async def trial_balance_extended(
    current_user: CurrentUser,
    db: DBSession,
    as_of: date | None = Query(None, description="Trial balance as of date (default: today)"),
) -> dict[str, Any]:
    as_of_date = as_of or date.today()

    # All active accounts with their journal line totals up to as_of
    result = await db.execute(
        select(
            Account.id,
            Account.code,
            Account.name,
            Account.account_type,
            func.coalesce(func.sum(JournalLine.debit), 0).label("total_debit"),
            func.coalesce(func.sum(JournalLine.credit), 0).label("total_credit"),
        )
        .outerjoin(JournalLine, JournalLine.account_id == Account.id)
        .outerjoin(
            JournalEntry,
            and_(
                JournalEntry.id == JournalLine.journal_entry_id,
                JournalEntry.status == "posted",
                JournalEntry.entry_date <= as_of_date,
            ),
        )
        .where(Account.is_active == True)  # noqa: E712
        .group_by(Account.id, Account.code, Account.name, Account.account_type)
        .order_by(Account.code.asc())
    )
    rows = result.all()

    accounts = []
    grand_debit = Decimal("0")
    grand_credit = Decimal("0")

    for row in rows:
        total_debit = Decimal(str(row.total_debit))
        total_credit = Decimal(str(row.total_credit))
        balance = total_debit - total_credit
        grand_debit += total_debit
        grand_credit += total_credit
        accounts.append({
            "account_id": str(row.id),
            "code": row.code,
            "name": row.name,
            "account_type": row.account_type,
            "total_debit": str(total_debit),
            "total_credit": str(total_credit),
            "balance": str(balance),
        })

    return {
        "as_of": str(as_of_date),
        "accounts": accounts,
        "grand_total_debit": str(grand_debit),
        "grand_total_credit": str(grand_credit),
        "is_balanced": grand_debit == grand_credit,
    }


# ── Aged Receivables ────────────────────────────────────────────────────────

@router.get("/reports/aged-receivables", summary="Aged receivables report")
async def aged_receivables(
    current_user: CurrentUser,
    db: DBSession,
    as_of: date | None = Query(None, description="As-of date (default: today)"),
) -> dict[str, Any]:
    as_of_date = as_of or date.today()

    # Unpaid sales invoices (sent or overdue)
    result = await db.execute(
        select(Invoice)
        .where(
            and_(
                Invoice.invoice_type == "sales",
                Invoice.status.in_(["sent", "overdue"]),
                Invoice.due_date <= as_of_date,
            )
        )
        .order_by(Invoice.due_date.asc())
    )
    invoices = result.scalars().all()

    buckets = {
        "current": [],       # not yet due
        "1_30": [],          # 1-30 days overdue
        "31_60": [],         # 31-60 days overdue
        "61_90": [],         # 61-90 days overdue
        "over_90": [],       # 90+ days overdue
    }
    totals = {k: Decimal("0") for k in buckets}

    # Also get invoices not yet due
    not_due_result = await db.execute(
        select(Invoice)
        .where(
            and_(
                Invoice.invoice_type == "sales",
                Invoice.status.in_(["sent", "overdue"]),
                Invoice.due_date > as_of_date,
            )
        )
        .order_by(Invoice.due_date.asc())
    )
    not_due_invoices = not_due_result.scalars().all()

    for inv in not_due_invoices:
        entry = {
            "invoice_id": str(inv.id),
            "invoice_number": inv.invoice_number,
            "customer_name": inv.customer_name,
            "due_date": str(inv.due_date),
            "total": str(inv.total),
            "currency": inv.currency,
            "days_overdue": 0,
        }
        buckets["current"].append(entry)
        totals["current"] += inv.total

    for inv in invoices:
        days_overdue = (as_of_date - inv.due_date).days
        entry = {
            "invoice_id": str(inv.id),
            "invoice_number": inv.invoice_number,
            "customer_name": inv.customer_name,
            "due_date": str(inv.due_date),
            "total": str(inv.total),
            "currency": inv.currency,
            "days_overdue": days_overdue,
        }

        if days_overdue <= 30:
            buckets["1_30"].append(entry)
            totals["1_30"] += inv.total
        elif days_overdue <= 60:
            buckets["31_60"].append(entry)
            totals["31_60"] += inv.total
        elif days_overdue <= 90:
            buckets["61_90"].append(entry)
            totals["61_90"] += inv.total
        else:
            buckets["over_90"].append(entry)
            totals["over_90"] += inv.total

    grand_total = sum(totals.values())

    return {
        "as_of": str(as_of_date),
        "buckets": {
            "current": {"invoices": buckets["current"], "total": str(totals["current"])},
            "1_30_days": {"invoices": buckets["1_30"], "total": str(totals["1_30"])},
            "31_60_days": {"invoices": buckets["31_60"], "total": str(totals["31_60"])},
            "61_90_days": {"invoices": buckets["61_90"], "total": str(totals["61_90"])},
            "over_90_days": {"invoices": buckets["over_90"], "total": str(totals["over_90"])},
        },
        "grand_total": str(grand_total),
    }


# ── Aged Payables ────────────────────────────────────────────────────────────

@router.get("/reports/aged-payables", summary="Aged payables report")
async def aged_payables(
    current_user: CurrentUser,
    db: DBSession,
    as_of: date | None = Query(None, description="As-of date (default: today)"),
) -> dict[str, Any]:
    as_of_date = as_of or date.today()

    # Unpaid vendor bills (draft, received, approved, overdue — not paid/cancelled)
    result = await db.execute(
        select(VendorBill)
        .where(
            and_(
                VendorBill.status.in_(["draft", "received", "approved", "overdue"]),
                VendorBill.due_date <= as_of_date,
            )
        )
        .order_by(VendorBill.due_date.asc())
    )
    bills = result.scalars().all()

    # Not yet due
    not_due_result = await db.execute(
        select(VendorBill)
        .where(
            and_(
                VendorBill.status.in_(["draft", "received", "approved", "overdue"]),
                VendorBill.due_date > as_of_date,
            )
        )
        .order_by(VendorBill.due_date.asc())
    )
    not_due_bills = not_due_result.scalars().all()

    buckets = {
        "current": [],
        "1_30": [],
        "31_60": [],
        "61_90": [],
        "over_90": [],
    }
    totals = {k: Decimal("0") for k in buckets}

    for bill in not_due_bills:
        entry = {
            "bill_id": str(bill.id),
            "bill_number": bill.bill_number,
            "vendor_name": bill.vendor_name,
            "due_date": str(bill.due_date),
            "total": str(bill.total),
            "currency": bill.currency,
            "days_overdue": 0,
        }
        buckets["current"].append(entry)
        totals["current"] += bill.total

    for bill in bills:
        days_overdue = (as_of_date - bill.due_date).days
        entry = {
            "bill_id": str(bill.id),
            "bill_number": bill.bill_number,
            "vendor_name": bill.vendor_name,
            "due_date": str(bill.due_date),
            "total": str(bill.total),
            "currency": bill.currency,
            "days_overdue": days_overdue,
        }

        if days_overdue <= 30:
            buckets["1_30"].append(entry)
            totals["1_30"] += bill.total
        elif days_overdue <= 60:
            buckets["31_60"].append(entry)
            totals["31_60"] += bill.total
        elif days_overdue <= 90:
            buckets["61_90"].append(entry)
            totals["61_90"] += bill.total
        else:
            buckets["over_90"].append(entry)
            totals["over_90"] += bill.total

    grand_total = sum(totals.values())

    return {
        "as_of": str(as_of_date),
        "buckets": {
            "current": {"bills": buckets["current"], "total": str(totals["current"])},
            "1_30_days": {"bills": buckets["1_30"], "total": str(totals["1_30"])},
            "31_60_days": {"bills": buckets["31_60"], "total": str(totals["31_60"])},
            "61_90_days": {"bills": buckets["61_90"], "total": str(totals["61_90"])},
            "over_90_days": {"bills": buckets["over_90"], "total": str(totals["over_90"])},
        },
        "grand_total": str(grand_total),
    }


# ── Dashboard KPIs ───────────────────────────────────────────────────────────

@router.get("/dashboard/kpis", summary="Financial KPIs for the dashboard")
async def dashboard_kpis(
    current_user: CurrentUser,
    db: DBSession,
    period_start: date | None = Query(None, description="Period start (default: first day of current month)"),
    period_end: date | None = Query(None, description="Period end (default: today)"),
) -> dict[str, Any]:
    today = date.today()
    start = period_start or today.replace(day=1)
    end = period_end or today

    # Revenue: sum of paid invoices (sales) in period
    revenue_result = await db.execute(
        select(
            func.coalesce(func.sum(Invoice.total), 0).label("total"),
            func.count(Invoice.id).label("count"),
        )
        .where(
            and_(
                Invoice.invoice_type == "sales",
                Invoice.status == "paid",
                Invoice.issue_date >= start,
                Invoice.issue_date <= end,
            )
        )
    )
    revenue_row = revenue_result.one()
    revenue = Decimal(str(revenue_row.total))
    invoices_paid = revenue_row.count

    # Expenses: sum of approved/reimbursed expenses in period
    expenses_result = await db.execute(
        select(
            func.coalesce(func.sum(Expense.amount), 0).label("total"),
            func.count(Expense.id).label("count"),
        )
        .where(
            and_(
                Expense.status.in_(["approved", "reimbursed"]),
                Expense.expense_date >= start,
                Expense.expense_date <= end,
            )
        )
    )
    expenses_row = expenses_result.one()
    total_expenses = Decimal(str(expenses_row.total))
    expense_count = expenses_row.count

    # Vendor bill payments in period
    vendor_paid_result = await db.execute(
        select(
            func.coalesce(func.sum(VendorBill.total), 0).label("total"),
        )
        .where(
            and_(
                VendorBill.status == "paid",
                VendorBill.issue_date >= start,
                VendorBill.issue_date <= end,
            )
        )
    )
    vendor_payments = Decimal(str(vendor_paid_result.scalar() or 0))

    total_outflows = total_expenses + vendor_payments
    profit = revenue - total_outflows

    # Cash position: total completed payments received minus total outgoing
    cash_in_result = await db.execute(
        select(func.coalesce(func.sum(Payment.amount), 0))
        .where(Payment.status == "completed")
    )
    total_cash_in = Decimal(str(cash_in_result.scalar() or 0))

    cash_out_result = await db.execute(
        select(func.coalesce(func.sum(VendorBill.total), 0))
        .where(VendorBill.status == "paid")
    )
    total_cash_out = Decimal(str(cash_out_result.scalar() or 0))

    expense_reimbursed_result = await db.execute(
        select(func.coalesce(func.sum(Expense.amount), 0))
        .where(Expense.status == "reimbursed")
    )
    total_reimbursed = Decimal(str(expense_reimbursed_result.scalar() or 0))

    cash_position = total_cash_in - total_cash_out - total_reimbursed

    # Receivables: sum of unpaid sales invoices
    receivables_result = await db.execute(
        select(func.coalesce(func.sum(Invoice.total), 0))
        .where(
            and_(
                Invoice.invoice_type == "sales",
                Invoice.status.in_(["sent", "overdue"]),
            )
        )
    )
    total_receivables = Decimal(str(receivables_result.scalar() or 0))

    # Payables: sum of unpaid vendor bills
    payables_result = await db.execute(
        select(func.coalesce(func.sum(VendorBill.total), 0))
        .where(
            VendorBill.status.in_(["draft", "received", "approved", "overdue"])
        )
    )
    total_payables = Decimal(str(payables_result.scalar() or 0))

    # Overdue counts
    overdue_invoices_result = await db.execute(
        select(func.count(Invoice.id))
        .where(
            and_(
                Invoice.invoice_type == "sales",
                Invoice.status.in_(["sent", "overdue"]),
                Invoice.due_date < today,
            )
        )
    )
    overdue_invoices = overdue_invoices_result.scalar() or 0

    overdue_bills_result = await db.execute(
        select(func.count(VendorBill.id))
        .where(
            and_(
                VendorBill.status.in_(["draft", "received", "approved", "overdue"]),
                VendorBill.due_date < today,
            )
        )
    )
    overdue_bills = overdue_bills_result.scalar() or 0

    # Pending expense approvals
    pending_expenses_result = await db.execute(
        select(func.count(Expense.id))
        .where(Expense.status == "submitted")
    )
    pending_expenses = pending_expenses_result.scalar() or 0

    return {
        "period": {"start_date": str(start), "end_date": str(end)},
        "revenue": str(revenue),
        "invoices_paid": invoices_paid,
        "total_expenses": str(total_expenses),
        "expense_count": expense_count,
        "vendor_payments": str(vendor_payments),
        "total_outflows": str(total_outflows),
        "profit": str(profit),
        "cash_position": str(cash_position),
        "total_receivables": str(total_receivables),
        "total_payables": str(total_payables),
        "overdue_invoices": overdue_invoices,
        "overdue_bills": overdue_bills,
        "pending_expense_approvals": pending_expenses,
    }


# ── XLSX Export ───────────────────────────────────────────────────────────────

from fastapi import Response as FastAPIResponse


@router.get("/reports/{report_type}/export-xlsx", summary="Export any finance report as XLSX")
async def export_report_xlsx(
    report_type: str,
    current_user: CurrentUser,
    db: DBSession,
    start_date: date = Query(default=None),
    end_date: date = Query(default=None),
) -> FastAPIResponse:
    """Export trial_balance, pl, balance_sheet, aged_ar, aged_ap, or kpis as XLSX."""
    import io

    import openpyxl
    from openpyxl.styles import Font, PatternFill

    today = date.today()
    start = start_date or date(today.year, 1, 1)
    end = end_date or today

    wb = openpyxl.Workbook()
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="51459D")

    def write_headers(sheet, headers: list[str]):
        for col, h in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

    if report_type == "trial_balance":
        ws.title = "Trial Balance"
        result = await db.execute(
            select(Account, func.coalesce(func.sum(JournalLine.debit), 0).label("total_debit"),
                   func.coalesce(func.sum(JournalLine.credit), 0).label("total_credit"))
            .outerjoin(JournalLine, JournalLine.account_id == Account.id)
            .group_by(Account.id).order_by(Account.code)
        )
        write_headers(ws, ["Account Code", "Account Name", "Type", "Total Debit", "Total Credit", "Net Balance"])
        for row_num, row in enumerate(result.all(), 2):
            net = float(row.total_debit) - float(row.total_credit)
            ws.append([row.Account.code, row.Account.name, row.Account.account_type,
                       float(row.total_debit), float(row.total_credit), net])

    elif report_type == "pl":
        ws.title = "Profit and Loss"
        rev_r = await db.execute(
            select(func.sum(Invoice.total)).where(Invoice.status == "paid", Invoice.invoice_type == "sales",
                                                   Invoice.issue_date.between(start, end))
        )
        exp_r = await db.execute(
            select(func.sum(Expense.amount)).where(Expense.status.in_(["approved", "reimbursed"]),
                                                    Expense.expense_date.between(start, end))
        )
        revenue = float(rev_r.scalar() or 0)
        expenses = float(exp_r.scalar() or 0)
        write_headers(ws, ["Metric", "Amount"])
        for label, val in [("Revenue", revenue), ("Operating Expenses", expenses),
                            ("Net Income", revenue - expenses)]:
            ws.append([label, val])

    elif report_type in ("aged_ar", "aged_ap"):
        ws.title = "Aged AR" if report_type == "aged_ar" else "Aged AP"
        write_headers(ws, ["Number", "Counterparty", "Amount", "Due Date", "Days Overdue", "Status"])
        if report_type == "aged_ar":
            result = await db.execute(
                select(Invoice).where(Invoice.status.in_(["sent", "overdue"])).order_by(Invoice.due_date)
            )
            items = result.scalars().all()
            for item in items:
                days = max(0, (today - item.due_date).days) if item.due_date else 0
                ws.append([item.invoice_number, item.customer_name, float(item.total or 0),
                           str(item.due_date), days, item.status])
        else:
            result = await db.execute(
                select(VendorBill).where(VendorBill.status.in_(["approved", "overdue"])).order_by(VendorBill.due_date)
            )
            items = result.scalars().all()
            for item in items:
                days = max(0, (today - item.due_date).days) if item.due_date else 0
                ws.append([item.bill_number, item.vendor_name, float(item.total or 0),
                           str(item.due_date), days, item.status])

    else:
        ws.title = report_type.replace("_", " ").title()
        write_headers(ws, ["Report Type", "Generated At"])
        ws.append([report_type, str(today)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{report_type}_{start}_{end}.xlsx"

    return FastAPIResponse(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Revenue Streams Report ────────────────────────────────────────────────────


@router.get("/reports/revenue-streams", summary="Revenue breakdown by account group, department, project")
async def revenue_streams_report(
    current_user: CurrentUser,
    db: DBSession,
    start_date: date = Query(default=None),
    end_date: date = Query(default=None),
) -> dict[str, Any]:
    today = date.today()
    start = start_date or date(today.year, 1, 1)
    end = end_date or today

    # Revenue by account (account type = revenue)
    by_account_r = await db.execute(
        select(Account.name.label("account"), Account.code.label("code"),
               func.coalesce(func.sum(JournalLine.credit), 0).label("amount"))
        .join(JournalLine, JournalLine.account_id == Account.id)
        .join(JournalEntry, JournalEntry.id == JournalLine.journal_entry_id)
        .where(Account.account_type == "revenue",
               JournalEntry.entry_date.between(start, end),
               JournalEntry.status == "posted")
        .group_by(Account.id, Account.name, Account.code)
        .order_by(func.sum(JournalLine.credit).desc())
    )
    by_account = [{"account": r.account, "code": r.code, "amount": float(r.amount)}
                  for r in by_account_r.all()]

    # Revenue from invoices (by customer)
    by_customer_r = await db.execute(
        select(Invoice.customer_name.label("customer"),
               func.count(Invoice.id).label("invoice_count"),
               func.sum(Invoice.total).label("revenue"))
        .where(Invoice.status == "paid", Invoice.invoice_type == "sales",
               Invoice.issue_date.between(start, end))
        .group_by(Invoice.customer_name)
        .order_by(func.sum(Invoice.total).desc())
        .limit(20)
    )
    by_customer = [{"customer": r.customer, "invoice_count": r.invoice_count, "revenue": float(r.revenue or 0)}
                   for r in by_customer_r.all()]

    # Monthly trend
    monthly_r = await db.execute(
        select(func.date_trunc("month", Invoice.issue_date).label("month"),
               func.sum(Invoice.total).label("revenue"))
        .where(Invoice.status == "paid", Invoice.invoice_type == "sales",
               Invoice.issue_date.between(start, end))
        .group_by(func.date_trunc("month", Invoice.issue_date))
        .order_by(func.date_trunc("month", Invoice.issue_date))
    )
    monthly_trend = [{"month": r.month.strftime("%Y-%m") if r.month else None,
                      "revenue": float(r.revenue or 0)}
                     for r in monthly_r.all()]

    total_revenue = sum(s["amount"] for s in by_account) or sum(c["revenue"] for c in by_customer)

    return {
        "period": {"start_date": str(start), "end_date": str(end)},
        "total_revenue": total_revenue,
        "by_account": by_account,
        "by_customer": by_customer,
        "monthly_trend": monthly_trend,
    }


# ── Project Job Costing Report ────────────────────────────────────────────────


@router.get("/reports/project-costing/{project_id}", summary="Project-level P&L: revenue, costs, overhead, margin")
async def project_costing_report(
    project_id: str,
    current_user: CurrentUser,
    db: DBSession,
) -> dict[str, Any]:
    import uuid as _uuid
    from app.models.finance import Budget

    pid = _uuid.UUID(project_id)

    # Direct expenses linked to this project
    exp_r = await db.execute(
        select(Expense.category,
               func.sum(Expense.amount).label("total"))
        .where(Expense.project_id == pid,
               Expense.status.in_(["approved", "reimbursed"]))
        .group_by(Expense.category)
    )
    expenses_by_category = [{"category": r.category, "amount": float(r.total or 0)}
                             for r in exp_r.all()]
    direct_costs = sum(e["amount"] for e in expenses_by_category)

    # Revenue from invoices linked to project (via description or project_id if available)
    # We look at paid invoices where description references the project
    rev_r = await db.execute(
        select(func.sum(Invoice.total)).where(
            Invoice.status == "paid",
            Invoice.invoice_type == "sales",
            Invoice.project_id == pid,
        )
    )
    revenue = float(rev_r.scalar() or 0)

    # Budget for project
    budget_r = await db.execute(
        select(Budget).where(Budget.project_id == pid).limit(1)
    )
    budget = budget_r.scalar_one_or_none()
    budget_amount = float(budget.total_amount if budget and hasattr(budget, "total_amount") else 0)

    # Overhead allocation (15% of direct costs — standard rate)
    overhead_rate = 0.15
    overhead = round(direct_costs * overhead_rate, 2)
    total_costs = direct_costs + overhead

    margin = revenue - total_costs
    margin_pct = (margin / revenue * 100) if revenue > 0 else 0

    return {
        "project_id": project_id,
        "revenue": revenue,
        "direct_costs": direct_costs,
        "expenses_by_category": expenses_by_category,
        "overhead_rate_pct": overhead_rate * 100,
        "overhead_allocated": overhead,
        "total_costs": total_costs,
        "gross_margin": margin,
        "gross_margin_pct": round(margin_pct, 2),
        "budget_amount": budget_amount,
        "budget_variance": budget_amount - total_costs if budget_amount else None,
        "status": "over_budget" if budget_amount and total_costs > budget_amount else "on_budget",
    }


# ── Compliance Calendar CRUD ──────────────────────────────────────────────────

from app.models.finance_ext import ComplianceEvent


@router.get("/compliance-events", summary="List compliance calendar events")
async def list_compliance_events(
    current_user: CurrentUser,
    db: DBSession,
    status: str | None = Query(default=None),
    category: str | None = Query(default=None),
    jurisdiction: str | None = Query(default=None),
) -> dict[str, Any]:
    from sqlalchemy import asc

    query = select(ComplianceEvent)
    if status:
        query = query.where(ComplianceEvent.status == status)
    if category:
        query = query.where(ComplianceEvent.category == category)
    if jurisdiction:
        query = query.where(ComplianceEvent.jurisdiction.ilike(f"%{jurisdiction}%"))
    query = query.order_by(asc(ComplianceEvent.due_date))

    result = await db.execute(query)
    events = result.scalars().all()

    return {
        "items": [
            {
                "id": str(e.id),
                "title": e.title,
                "description": e.description,
                "category": e.category,
                "jurisdiction": e.jurisdiction,
                "due_date": str(e.due_date),
                "status": e.status,
                "recurrence": e.recurrence,
                "reminder_days": e.reminder_days,
                "assigned_to_id": str(e.assigned_to_id) if e.assigned_to_id else None,
            }
            for e in events
        ],
        "total": len(events),
    }


from pydantic import BaseModel as _BaseModel


class ComplianceEventCreate(_BaseModel):
    title: str
    description: str | None = None
    category: str  # tax_filing, regulatory, payroll, annual_return, other
    jurisdiction: str | None = None
    due_date: date
    recurrence: str | None = None  # monthly, quarterly, annually
    reminder_days: list[int] = [30, 7, 1]
    assigned_to_id: str | None = None


class ComplianceEventIn(_BaseModel):
    title: str
    description: str | None = None
    category: str
    jurisdiction: str | None = None
    due_date: date
    recurrence: str | None = None
    reminder_days: list[int] = [30, 7, 1]
    assigned_to_id: str | None = None


@router.post("/compliance-events", summary="Create a compliance calendar event")
async def create_compliance_event(
    payload: ComplianceEventIn,
    current_user: CurrentUser,
    db: DBSession,
) -> dict[str, Any]:
    import uuid as _uuid

    evt = ComplianceEvent(
        title=payload.title,
        description=payload.description,
        category=payload.category,
        jurisdiction=payload.jurisdiction,
        due_date=payload.due_date,
        recurrence=payload.recurrence,
        reminder_days=payload.reminder_days,
        assigned_to_id=_uuid.UUID(payload.assigned_to_id) if payload.assigned_to_id else None,
        status="pending",
        created_by_id=current_user.id,
    )
    db.add(evt)
    await db.commit()
    await db.refresh(evt)
    return {"id": str(evt.id), "title": evt.title, "due_date": str(evt.due_date), "status": evt.status}


@router.put("/compliance-events/{event_id}", summary="Update a compliance event")
async def update_compliance_event(
    event_id: str,
    payload: ComplianceEventIn,
    current_user: CurrentUser,
    db: DBSession,
) -> dict[str, Any]:
    import uuid as _uuid

    result = await db.execute(select(ComplianceEvent).where(ComplianceEvent.id == _uuid.UUID(event_id)))
    evt = result.scalar_one_or_none()
    if not evt:
        from fastapi import HTTPException as _HTTPException
        raise _HTTPException(status_code=404, detail="Compliance event not found")

    for field, val in payload.model_dump(exclude_none=True).items():
        if field == "assigned_to_id" and val:
            setattr(evt, field, _uuid.UUID(val))
        else:
            setattr(evt, field, val)

    await db.commit()
    return {"id": str(evt.id), "title": evt.title, "status": evt.status}


@router.patch("/compliance-events/{event_id}/complete", summary="Mark compliance event as completed")
async def complete_compliance_event(
    event_id: str,
    current_user: CurrentUser,
    db: DBSession,
) -> dict[str, Any]:
    import uuid as _uuid

    result = await db.execute(select(ComplianceEvent).where(ComplianceEvent.id == _uuid.UUID(event_id)))
    evt = result.scalar_one_or_none()
    if not evt:
        from fastapi import HTTPException as _HTTPException
        raise _HTTPException(status_code=404, detail="Compliance event not found")

    evt.status = "completed"
    await db.commit()
    return {"id": str(evt.id), "status": "completed"}


@router.delete("/compliance-events/{event_id}", summary="Delete a compliance event")
async def delete_compliance_event(
    event_id: str,
    current_user: CurrentUser,
    db: DBSession,
) -> dict[str, Any]:
    import uuid as _uuid

    result = await db.execute(select(ComplianceEvent).where(ComplianceEvent.id == _uuid.UUID(event_id)))
    evt = result.scalar_one_or_none()
    if not evt:
        from fastapi import HTTPException as _HTTPException
        raise _HTTPException(status_code=404, detail="Compliance event not found")

    await db.delete(evt)
    await db.commit()
    return {"deleted": event_id}


# ── FX Revaluation API ────────────────────────────────────────────────────────

from app.models.finance_ext import FXRevaluationEntry


@router.get("/fx-revaluations", summary="List FX revaluation entries")
async def list_fx_revaluations(
    current_user: CurrentUser,
    db: DBSession,
    period: str | None = Query(default=None, description="Filter by period e.g. 2026-01"),
) -> dict[str, Any]:
    query = select(FXRevaluationEntry).order_by(FXRevaluationEntry.revaluation_date.desc())
    if period:
        query = query.where(FXRevaluationEntry.period == period)
    result = await db.execute(query)
    entries = result.scalars().all()

    return {
        "items": [
            {
                "id": str(e.id),
                "account_id": str(e.account_id),
                "period": e.period,
                "currency": e.currency,
                "original_balance": float(e.original_balance or 0),
                "revalued_balance": float(e.revalued_balance or 0),
                "fx_rate": float(e.fx_rate or 1),
                "unrealized_gain_loss": float(e.unrealized_gain_loss or 0),
                "is_realized": e.is_realized,
                "revaluation_date": str(e.revaluation_date),
                "journal_entry_id": str(e.journal_entry_id) if e.journal_entry_id else None,
            }
            for e in entries
        ],
        "total": len(entries),
        "total_unrealized_gl": sum(float(e.unrealized_gain_loss or 0) for e in entries),
    }


@router.post("/fx-revaluations/run", summary="Manually trigger FX revaluation for current period")
async def trigger_fx_revaluation(
    current_user: CurrentUser,
    db: DBSession,
) -> dict[str, Any]:
    from app.tasks.celery_app import run_fx_revaluation
    task = run_fx_revaluation.delay()
    return {"task_id": task.id, "status": "queued", "message": "FX revaluation queued as background task"}
